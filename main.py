import os
import sys
import re
import copy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MUNICIPALITIES = {'北京市', '上海市', '天津市', '重庆市'}

# ======================================================
# 地址去重清洗
# ======================================================

def clean_duplicate_address(addr):
    if not isinstance(addr, str) or str(addr).strip() in ['nan', '']:
        return ""

    addr = str(addr).strip()
    prov, city = "", ""

    if addr[:3] in MUNICIPALITIES:
        prov = addr[:3]
    else:
        match = re.match(
            r'^(.{2,8}?(?:省|自治区))(.{2,8}?(?:市|自治州|地区|盟))?',
            addr
        )
        if match:
            prov = match.group(1)
            city = match.group(2) or ""

    prefix = prov + city

    if prefix:
        second_idx = addr.find(prefix, len(prov))
        if second_idx > 0 and second_idx <= 25:
            middle = addr[len(prefix):second_idx]
            if re.match(r'^.{1,10}?(?:区|县|镇|乡|街道|旗|村)$', middle):
                match_second = re.match(
                    r'^' + re.escape(prefix) +
                    r'.{1,10}?(?:区|县|镇|乡|街道|旗|村)',
                    addr[second_idx:]
                )
                if match_second:
                    addr = (
                        addr[:second_idx] +
                        addr[second_idx + len(match_second.group(0)):]
                    )
                else:
                    addr = (
                        addr[:second_idx] +
                        addr[second_idx + len(prefix):]
                    )
            elif middle == "":
                addr = addr[second_idx:]

    n = len(addr)
    changed = True
    while changed:
        changed = False
        for l in range(n // 2, 2, -1):
            sub = addr[:l]
            if addr[l:].startswith(sub):
                addr = addr[l:]
                n = len(addr)
                changed = True
                break

    changed = True
    while changed:
        changed = False
        n = len(addr)
        for i in range(n):
            for l in range(3, (n - i) // 2 + 1):
                sub = addr[i:i+l]
                if addr[i+l:i+2*l] == sub:
                    # 🚀 修复核心：正确保留前面匹配到的子串(i+l)，跳过后面的重复子串(i+2*l)
                    addr = addr[:i+l] + addr[i+2*l:]
                    changed = True
                    break
            if changed:
                break

    return addr

# ======================================================
# 地址解析
# ======================================================

def parse_address(addr):
    addr = str(addr).strip()

    # ---- 特例前置或后置处理均可，这里在标准逻辑跑完后进行修正 ----
    if addr[:3] in MUNICIPALITIES:
        province = addr[:3]
        pe = 3
    elif '自治区' in addr:
        pe = addr.index('自治区') + 3
        province = addr[:pe]
    elif '省' in addr:
        pe = addr.index('省') + 1
        province = addr[:pe]
    else:
        province = addr[:3]
        pe = 3

    try:
        ce = addr.index('市', pe) + 1
    except:
        ce = pe

    city = addr[pe:ce]
    if not city:
        city = province

    de = len(addr)
    for ch in ['区', '县', '镇', '乡', '旗', '街道']:
        try:
            pos = addr.index(ch, ce) + len(ch)
            if pos < de:
                de = pos
        except:
            pass

    try:
        pos = addr.index('市', ce) + 1
        if pos < de:
            de = pos
    except:
        pass

    district = addr[ce:de]
    detail = addr[de:]

    for prefix in [addr[:de], city + district, district]:
        if prefix and detail.startswith(prefix):
            detail = detail[len(prefix):]
            break

    if city in ['中山市', '东莞市', '嘉峪关市', '儋州市'] and district != city:
        if district:
            detail = district + detail
        district = city

    # ======================================================
    # 新增：针对特定地址的硬编码修正逻辑
    # ======================================================
    if "东荆河路175号（武汉美安物流园）" in addr:
        district = "蔡甸区"
        detail = "东荆河路175号（武汉美安物流园）"
    elif "花溪街与浔江东路交汇西北角中原物流园" in addr:
        district = "经开县"
        detail = "花溪街与浔江东路交汇西北角中原物流园7号库U2单元"

    return province, city, district, detail

# ======================================================
# 路径
# ======================================================
if getattr(sys, 'frozen', False):
    CURRENT_DIR = os.path.dirname(sys.executable)
else:
    CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))

path_12 = os.path.join(CURRENT_DIR, "12.xlsx")
path_123 = os.path.join(CURRENT_DIR, "123.xlsx")

if not os.path.exists(path_12) or not os.path.exists(path_123):
    print("\n【错误】缺少xlsx文件！")
    input("\n按回车退出...")
    sys.exit()

output_path = os.path.join(CURRENT_DIR, "123_output.xlsx")

# ======================================================
# 第一步：读取12数据
# ======================================================
print("=================== 第一步：读取12数据 ===================")
df_12 = pd.read_excel(path_12, dtype=str)
df_12.columns = df_12.columns.str.strip()
df_12 = df_12.fillna('')

if '收货地址' in df_12.columns:
    df_12['收货地址'] = df_12['收货地址'].str.replace('广西壮族自治区', '广西省')
    df_12['收货地址'] = df_12['收货地址'].apply(clean_duplicate_address)

# ======================================================
# 第二步：重构123模板
# ======================================================
print("\n=================== 第二步：重构123模板 ===================")
wb = load_workbook(path_123)
ws = wb.active

current_headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(1, col).value
    if val:
        current_headers[str(val).strip()] = col

anchor_idx = current_headers.get('收货地址（复制）') or current_headers.get('收货地址')

if anchor_idx:
    ws.cell(1, anchor_idx).value = '收货地址（复制）'
    if '收货省份' not in current_headers:
        old_col_props = {}
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            if letter in ws.column_dimensions:
                dim = ws.column_dimensions[letter]
                old_col_props[col_idx] = {
                    'width': dim.width,
                    'hidden': dim.hidden
                }
            else:
                old_col_props[col_idx] = {
                    'width': None,
                    'hidden': False
                }

        ws.insert_cols(anchor_idx, 4)
        ws.cell(1, anchor_idx).value = '收货省份'
        ws.cell(1, anchor_idx + 1).value = '收货城市'
        ws.cell(1, anchor_idx + 2).value = '收货区县'
        ws.cell(1, anchor_idx + 3).value = '收货地址'

        ws.column_dimensions.clear()
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            if col_idx < anchor_idx:
                props = old_col_props.get(col_idx)
            elif anchor_idx <= col_idx < anchor_idx + 4:
                w = 35 if (col_idx - anchor_idx == 3) else 15
                props = {
                    'width': w,
                    'hidden': False
                }
            else:
                props = old_col_props.get(col_idx - 4)

            if props:
                if props['width'] is not None:
                    ws.column_dimensions[letter].width = props['width']
                if props['hidden']:
                    ws.column_dimensions[letter].hidden = True

headers_123 = {}
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    if header:
        headers_123[str(header).strip()] = col

target_data_count = len(df_12)
excel_data_rows = ws.max_row - 1

if excel_data_rows > target_data_count:
    ws.delete_rows(
        idx=target_data_count + 2,
        amount=excel_data_rows - target_data_count
    )

# ======================================================
# 第三步：复制主干字段
# ======================================================
print("\n=================== 第三步：复制主干字段 ===================")
copy_map = {
    '订单号': '其它出库业务单号',
    '收货人': '收货人',
    '收货电话': '收货电话',
    '收货地址': '收货地址（复制）',
    'SKU采购总⾦额（含税）': '单价',
    '采购数量（采购单位）': '数量',
    'SKU编码': 'SKU编码',
}

for source_col, target_col in copy_map.items():
    if source_col in df_12.columns and target_col in headers_123:
        target_excel_col = headers_123[target_col]
        for i in range(target_data_count):
            val = df_12.iloc[i][source_col]
            if target_col in ['SKU编码', '其它出库业务单号', '收货电话']:
                val = str(val).strip()
                if val.endswith('.0'):
                    val = val[:-2]
            ws.cell(i + 2, target_excel_col).value = val

# ======================================================
# 第四步：地址解析与备注
# ======================================================
print("\n=================== 第四步：地址解析与备注 ===================")
if '收货地址' in df_12.columns:
    parsed_data = [
        parse_address(addr)
        if str(addr).strip()
        else ("", "", "", "")
        for addr in df_12['收货地址']
    ]

    parsed_df = pd.DataFrame(
        parsed_data,
        columns=['收货省份', '收货城市', '收货区县', '收货地址']
    )

    # 修改：为了使备注中的地址包含刚刚修正后的“区/县”，
    # 我们用‘收货省份+收货城市+收货区县+收货地址’重新组合成修正后的完整地址
    full_corrected_addresses = (
        parsed_df['收货省份'] + 
        parsed_df['收货城市'] + 
        parsed_df['收货区县'] + 
        parsed_df['收货地址']
    ).str.strip()

    parsed_df['备注'] = (
        df_12['收货人']
        .astype(str)
        .fillna('')
        .str.replace('nan', '')
        + " " +
        df_12['收货电话']
        .astype(str)
        .fillna('')
        .str.replace('nan', '')
        + " " +
        full_corrected_addresses
    ).str.strip()

    for field in parsed_df.columns:
        if field not in headers_123:
            new_col = ws.max_column + 1
            ws.cell(1, new_col).value = field
            headers_123[field] = new_col

        target_excel_col = headers_123[field]
        for i in range(target_data_count):
            ws.cell(
                i + 2,
                target_excel_col
            ).value = parsed_df.iloc[i][field]

# ======================================================
# 统一字体与对齐
# ======================================================
print("\n=================== 统一格式 ===================")
for col in range(1, ws.max_column + 1):
    sample_cell = ws.cell(2, col)
    sample_font = sample_cell.font
    sample_alignment = sample_cell.alignment

    for r_idx in range(3, target_data_count + 2):
        cell = ws.cell(r_idx, col)
        if sample_font:
            cell.font = copy.copy(sample_font)
        if sample_alignment:
            cell.alignment = copy.copy(sample_alignment)

# ======================================================
# sheet名称
# ======================================================
ws.title = "总表"

# ======================================================
# 保存
# ======================================================
print("\n=================== 保存文件 ===================")
wb.save(output_path)
print("\n【✨ 全部完成 ✨】")
input("\n按回车退出...")
